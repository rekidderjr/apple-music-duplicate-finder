#!/usr/bin/env python3
"""
Apple Music Library Duplicate Finder

This script analyzes an Apple Music Library.XML file to identify duplicate tracks
based on metadata and file paths.
"""

import os
import sys
import xml.etree.ElementTree as ET
import plistlib
import datetime
import json
from collections import defaultdict
import argparse

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Find duplicate tracks in Apple Music Library')
    parser.add_argument('--input', '-i', default='data/Library.xml',
                        help='Path to Library.xml file (default: data/Library.xml)')
    parser.add_argument('--output', '-o', default='output',
                        help='Directory for output reports (default: output)')
    parser.add_argument('--allowlist', '-a', default='output/allowlist.json',
                        help='Path to allowlist file (default: output/allowlist.json)')
    return parser.parse_args()

def ensure_directory_exists(directory):
    """Create directory if it doesn't exist."""
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Created directory: {directory}")

def load_allowlist(allowlist_path):
    """Load the allowlist of duplicates to ignore."""
    if os.path.exists(allowlist_path):
        try:
            with open(allowlist_path, 'r') as f:
                return json.load(f)
        except json.JSONDecodeError:
            print(f"Warning: Allowlist file {allowlist_path} is not valid JSON. Creating a new one.")
            return {'metadata_duplicates': [], 'location_duplicates': []}
    else:
        print(f"Allowlist file not found at {allowlist_path}. Creating a new one.")
        return {'metadata_duplicates': [], 'location_duplicates': []}

def save_allowlist(allowlist, allowlist_path):
    """Save the allowlist to a file."""
    with open(allowlist_path, 'w') as f:
        json.dump(allowlist, f, indent=2)
    print(f"Allowlist saved to {allowlist_path}")

def load_library(file_path):
    """Load and parse the Apple Music Library.XML file."""
    if not os.path.exists(file_path):
        print(f"Error: Library file not found at {file_path}")
        print("Please export your Apple Music Library and place it in the specified location.")
        sys.exit(1)
    
    print(f"Loading library from {file_path}...")
    try:
        with open(file_path, 'rb') as file:
            library = plistlib.load(file)
        return library
    except Exception as e:
        print(f"Error loading library: {e}")
        sys.exit(1)

def extract_tracks(library):
    """Extract track information from the library."""
    if 'Tracks' not in library:
        print("Error: No tracks found in the library file.")
        sys.exit(1)
    
    return library['Tracks']

def find_duplicates_by_metadata(tracks, allowlist=None):
    """Find tracks with identical metadata but different file paths."""
    # Group by a combination of metadata fields
    metadata_groups = defaultdict(list)
    
    for track_id, track in tracks.items():
        # Skip tracks without location
        if 'Location' not in track:
            continue
        
        # Extract file extension from location
        location = track.get('Location', '')
        file_extension = os.path.splitext(location)[1].lower()
        
        # Create a metadata key using relevant fields including file extension
        metadata_key = (
            track.get('Name', ''),
            track.get('Artist', ''),
            track.get('Album', ''),
            track.get('Total Time', 0),
            file_extension  # Include file extension in the key
        )
        
        metadata_groups[metadata_key].append({
            'Track ID': track_id,
            'Name': track.get('Name', 'Unknown'),
            'Artist': track.get('Artist', 'Unknown'),
            'Album': track.get('Album', 'Unknown'),
            'Location': track.get('Location', ''),
            'Play Count': track.get('Play Count', 0),
            'Date Added': track.get('Date Added', ''),
            'File Extension': file_extension
        })
    
    # Filter for groups with more than one track
    duplicates = {key: tracks for key, tracks in metadata_groups.items() if len(tracks) > 1}
    
    # Filter out allowed duplicates
    if allowlist and 'metadata_duplicates' in allowlist:
        filtered_duplicates = {}
        for key, tracks in duplicates.items():
            # Check if this group is in the allowlist
            track_ids = sorted([track['Track ID'] for track in tracks])
            if track_ids not in allowlist['metadata_duplicates']:
                filtered_duplicates[key] = tracks
        return filtered_duplicates
    
    return duplicates

def find_duplicates_by_location(tracks, allowlist=None):
    """Find multiple entries pointing to the same file."""
    location_groups = defaultdict(list)
    
    for track_id, track in tracks.items():
        if 'Location' not in track:
            continue
        
        location = track['Location']
        location_groups[location].append({
            'Track ID': track_id,
            'Name': track.get('Name', 'Unknown'),
            'Artist': track.get('Artist', 'Unknown'),
            'Album': track.get('Album', 'Unknown'),
            'Play Count': track.get('Play Count', 0),
            'Date Added': track.get('Date Added', '')
        })
    
    # Filter for groups with more than one track
    duplicates = {loc: tracks for loc, tracks in location_groups.items() if len(tracks) > 1}
    
    # Filter out allowed duplicates
    if allowlist and 'location_duplicates' in allowlist:
        filtered_duplicates = {}
        for location, tracks in duplicates.items():
            # Check if this group is in the allowlist
            track_ids = sorted([track['Track ID'] for track in tracks])
            if track_ids not in allowlist['location_duplicates']:
                filtered_duplicates[location] = tracks
        return filtered_duplicates
    
    return duplicates

def generate_report(metadata_duplicates, location_duplicates, output_dir):
    """Generate reports for duplicate tracks."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Report for metadata duplicates
    metadata_report = {
        'report_type': 'Duplicate Tracks with Different Locations',
        'generated_at': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'total_duplicate_groups': len(metadata_duplicates),
        'duplicate_groups': []
    }
    
    # CSV data for metadata duplicates
    metadata_csv_rows = []
    metadata_csv_rows.append(['Group ID', 'Track ID', 'Name', 'Artist', 'Album', 
                             'Duration (ms)', 'File Extension', 'Play Count', 
                             'Date Added', 'Location'])
    
    group_id = 1
    for key, tracks in metadata_duplicates.items():
        # Convert any datetime objects in tracks to strings
        processed_tracks = []
        for track in tracks:
            processed_track = {}
            for k, v in track.items():
                if isinstance(v, datetime.datetime):
                    processed_track[k] = v.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    processed_track[k] = v
            processed_tracks.append(processed_track)
            
            # Add to CSV data
            metadata_csv_rows.append([
                group_id,
                track['Track ID'],
                track['Name'],
                track['Artist'],
                track['Album'],
                key[3],  # Duration from the key tuple
                track.get('File Extension', ''),
                track.get('Play Count', 0),
                track.get('Date Added', ''),
                track['Location']
            ])
            
        metadata_report['duplicate_groups'].append({
            'name': key[0],
            'artist': key[1],
            'album': key[2],
            'duration': key[3],
            'file_extension': key[4],  # Add file extension to the report
            'tracks': processed_tracks
        })
        group_id += 1
    
    # Report for location duplicates
    location_report = {
        'report_type': 'Multiple Entries with Same Location',
        'generated_at': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'total_duplicate_groups': len(location_duplicates),
        'duplicate_groups': []
    }
    
    # CSV data for location duplicates
    location_csv_rows = []
    location_csv_rows.append(['Group ID', 'Track ID', 'Name', 'Artist', 'Album', 
                             'Play Count', 'Date Added', 'Location'])
    
    group_id = 1
    for location, tracks in location_duplicates.items():
        # Convert any datetime objects in tracks to strings
        processed_tracks = []
        for track in tracks:
            processed_track = {}
            for k, v in track.items():
                if isinstance(v, datetime.datetime):
                    processed_track[k] = v.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    processed_track[k] = v
            processed_tracks.append(processed_track)
            
            # Add to CSV data
            location_csv_rows.append([
                group_id,
                track['Track ID'],
                track.get('Name', 'Unknown'),
                track.get('Artist', 'Unknown'),
                track.get('Album', 'Unknown'),
                track.get('Play Count', 0),
                track.get('Date Added', ''),
                location
            ])
            
        location_report['duplicate_groups'].append({
            'location': location,
            'tracks': processed_tracks
        })
        group_id += 1
    
    # Write reports to files - without timestamps
    metadata_file = os.path.join(output_dir, "metadata_duplicates.json")
    location_file = os.path.join(output_dir, "location_duplicates.json")
    excel_file = os.path.join(output_dir, "duplicates.xlsx")
    
    # Create backup of previous files if they exist
    for file_path in [metadata_file, location_file, excel_file]:
        if os.path.exists(file_path):
            backup_file = f"{file_path}.bak"
            try:
                os.replace(file_path, backup_file)
                print(f"Created backup of previous file: {backup_file}")
            except Exception as e:
                print(f"Warning: Could not create backup of {file_path}: {e}")
    
    # Write JSON reports
    with open(metadata_file, 'w') as f:
        json.dump(metadata_report, f, indent=2)
    
    with open(location_file, 'w') as f:
        json.dump(location_report, f, indent=2)
    
    # Write Excel report with multiple sheets
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet as the first sheet
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Group metadata duplicates by file extension
        extension_groups = {}
        for row_data in metadata_csv_rows[1:]:  # Skip header row
            file_extension = row_data[6]  # File Extension column
            if file_extension not in extension_groups:
                extension_groups[file_extension] = []
            extension_groups[file_extension].append(row_data)
        
        # Add headers with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Create a sheet for each file extension
        for file_extension, rows in extension_groups.items():
            # Clean up extension name for sheet title
            sheet_name = file_extension.strip().lstrip('.').upper() if file_extension else "Unknown"
            if not sheet_name:
                sheet_name = "Unknown"
            
            # Ensure sheet name is valid and not too long
            sheet_name = sheet_name[:31]  # Excel limits sheet names to 31 chars
            
            # Create sheet
            ext_sheet = wb.create_sheet(title=sheet_name)
            
            # Add headers
            for col_idx, header in enumerate(metadata_csv_rows[0], 1):
                cell = ext_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ext_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Create "All Metadata Duplicates" sheet
        all_metadata_sheet = wb.create_sheet(title="All Metadata Duplicates")
        
        # Add headers
        for col_idx, header in enumerate(metadata_csv_rows[0], 1):
            cell = all_metadata_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows
        for row_idx, row_data in enumerate(metadata_csv_rows[1:], 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                all_metadata_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Create location duplicates sheet
        if location_csv_rows and len(location_csv_rows) > 1:  # If we have data beyond headers
            location_sheet = wb.create_sheet(title="Location Duplicates")
            
            # Add headers with formatting
            for col_idx, header in enumerate(location_csv_rows[0], 1):
                cell = location_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add data rows
            for row_idx, row_data in enumerate(location_csv_rows[1:], 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    location_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Add summary information to summary sheet
        summary_sheet.cell(row=1, column=1, value="Apple Music Library Duplicate Analysis")
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        summary_sheet.cell(row=3, column=1, value="Analysis Date:")
        summary_sheet.cell(row=3, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        summary_sheet.cell(row=5, column=1, value="Duplicate Tracks with Different Locations:")
        summary_sheet.cell(row=5, column=2, value=len(metadata_duplicates))
        
        summary_sheet.cell(row=6, column=1, value="Multiple Entries with Same Location:")
        summary_sheet.cell(row=6, column=2, value=len(location_duplicates))
        
        # Add file extension breakdown
        summary_sheet.cell(row=8, column=1, value="Breakdown by File Extension:")
        row_num = 9
        for ext, rows in sorted(extension_groups.items()):
            ext_name = ext if ext else "Unknown"
            summary_sheet.cell(row=row_num, column=1, value=ext_name)
            summary_sheet.cell(row=row_num, column=2, value=len(rows))
            row_num += 1
        
        # Add instructions
        summary_sheet.cell(row=row_num + 1, column=1, value="Instructions:")
        summary_sheet.cell(row=row_num + 2, column=1, 
                          value="• Use the tabs at the bottom to navigate between different file types")
        summary_sheet.cell(row=row_num + 3, column=1, 
                          value="• Each tab contains duplicates for a specific file extension")
        summary_sheet.cell(row=row_num + 4, column=1, 
                          value="• 'All Metadata Duplicates' shows all tracks with identical metadata")
        summary_sheet.cell(row=row_num + 5, column=1, 
                          value="• 'Location Duplicates' shows multiple entries pointing to the same file")
        
        # Adjust column widths for better readability
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                # Cap at 50 for very long paths
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Save the workbook
        wb.save(excel_file)
        excel_created = True
    except ImportError:
        print("Warning: openpyxl not installed. Falling back to CSV format.")
        excel_created = False
        
        # Write CSV reports as fallback
        metadata_csv_file = os.path.join(output_dir, f"metadata_duplicates_{timestamp}.csv")
        location_csv_file = os.path.join(output_dir, f"location_duplicates_{timestamp}.csv")
        
        import csv
        with open(metadata_csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(metadata_csv_rows)
        
        with open(location_csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(location_csv_rows)
    
    # Generate summary report
    summary_file = os.path.join(output_dir, "summary.txt")
    with open(summary_file, 'w') as f:
        f.write("Apple Music Library Duplicate Analysis\n")
        f.write("====================================\n\n")
        f.write(f"Analysis performed at: {datetime.datetime.now().isoformat()}\n\n")
        
        f.write("Duplicate Tracks with Different Locations\n")
        f.write("---------------------------------------\n")
        f.write(f"Found {len(metadata_duplicates)} groups of tracks with identical metadata ")
        f.write("but different locations.\n")
        f.write(f"Detailed reports: \n")
        f.write(f"  - JSON: {os.path.basename(metadata_file)}\n")
        if excel_created:
            f.write(f"  - Excel: {os.path.basename(excel_file)} (Metadata Duplicates sheet)\n")
        else:
            f.write(f"  - CSV: {os.path.basename(metadata_csv_file)}\n")
        f.write("\n")
        
        f.write("Multiple Entries with Same Location\n")
        f.write("----------------------------------\n")
        f.write(f"Found {len(location_duplicates)} instances of multiple entries ")
        f.write("pointing to the same file.\n")
        f.write(f"Detailed reports: \n")
        f.write(f"  - JSON: {os.path.basename(location_file)}\n")
        if excel_created:
            f.write(f"  - Excel: {os.path.basename(excel_file)} (Location Duplicates sheet)\n")
        else:
            f.write(f"  - CSV: {os.path.basename(location_csv_file)}\n")
    
    return metadata_file, location_file, summary_file

def main():
    """Main function to run the analysis."""
    args = parse_arguments()
    
    # Ensure output directory exists
    ensure_directory_exists(args.output)
    
    # Load allowlist
    allowlist = load_allowlist(args.allowlist)
    
    # Load and parse the library
    library = load_library(args.input)
    tracks = extract_tracks(library)
    
    print(f"Analyzing {len(tracks)} tracks...")
    
    # Find duplicates, filtering out allowed duplicates
    metadata_duplicates = find_duplicates_by_metadata(tracks, allowlist)
    location_duplicates = find_duplicates_by_location(tracks, allowlist)
    
    # Generate reports
    metadata_file, location_file, summary_file = generate_report(
        metadata_duplicates, location_duplicates, args.output
    )
    
    # Get the path for the Excel file
    excel_file = os.path.join(args.output, "duplicates.xlsx")
    
    # Print summary
    print("\nAnalysis complete!")
    print(f"Found {len(metadata_duplicates)} groups of tracks with identical metadata but different locations.")
    print(f"Found {len(location_duplicates)} instances of multiple entries pointing to the same file.")
    print(f"\nReports saved to:")
    print(f"  - {summary_file}")
    print(f"  - {metadata_file}")
    print(f"  - {location_file}")
    print(f"  - {excel_file}")
    print("\nTip: Open the Excel file to view all duplicates in a single workbook with multiple sheets.")

if __name__ == "__main__":
    main()
